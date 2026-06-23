# -*- coding: utf-8 -*-
"""验收用例.md(工程版 Gherkin) → 策划版验收清单 xlsx，可选真值代入。

把 `-05验收用例.md`(场景挂 R-编号、假设/当/那么 句式)自动翻成策划/QA
"拿着游戏逐条勾"的清单:说明页 + 测试清单页(下拉 通过/不通过/待测、
冻结表头、筛选、进度统计公式)。**从 Gherkin 自动抽用例,不手列。**

真值代入(--config):断言里的 `表名[主键].字段`(如 `levelTable[50].value`)
查配置表换成真值(`500`),并把来源记到「来源字段」列。这样策划拿真数字
直接核对游戏,同时**反向校验工程版/规则读表取值对不对**(配置↔规则一致性)。

护栏(绝不臆造):
  - 单主键且数字(`levelTable[50]`)→ 查 col0/id 列代入真值。
  - 多主键 / 枚举名键(`starTable[B,火,0]`)→ 标 `[需手填]`,**不猜**。
  - 查不到(表/字段/行缺)→ 标 `[查不到]`,保留原始断言文本。
  - --loc 给 LocalizationText 时,NameId/DescId 这类文本字段额外附中文。

用法:
  python gherkin_to_checklist.py <验收用例.md> [out.xlsx] [--config <配置目录>] [--loc <LocalizationText.xlsx>]
依赖 openpyxl(写侧);--config 时复用同目录 xlsx_dump.py 解配置表。
"""
import sys, os, re, datetime, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import xlsx_dump   # 复用 zipfile+xml 解析(openpyxl 读国产 xlsx 会报错)
import zipfile
from config_index import build_index, load_enums, load_keymap, lookup, REF_RE   # 共享索引/解析层

FONT = "微软雅黑"
HEAD_FILL  = PatternFill("solid", fgColor="2F7D18")
GRP_FILL   = PatternFill("solid", fgColor="EAF4E0")
TITLE_FILL = PatternFill("solid", fgColor="1D3A17")
thin = Side(style="thin", color="C9D6BD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

GIVEN = ("假设", "假定", "前提", "前置")
WHEN  = ("当",)
THEN  = ("那么",)
CONT  = ("并且", "而且", "但是", "并", "和")

# ---------- Gherkin 解析 ----------
def extract_r(text):
    m = re.search(r"[（(]\s*(R-[A-Za-z0-9\-/ ]+?)\s*[)）]\s*$", text)
    return (text[:m.start()].strip(), m.group(1).strip()) if m else (text.strip(), "")

def parse(md_text):
    feature, cases, cur, last = "", [], None, None
    def flush():
        if cur and (cur["given"] or cur["when"] or cur["then"] or cur["name"]):
            cases.append(cur)
    for raw in md_text.splitlines():
        line = raw.strip()
        if not line: continue
        m = re.match(r"^#*\s*功能\s*[:：]\s*(.+)$", line)
        if m: feature = extract_r(m.group(1))[0]; continue
        m = re.match(r"^场景\s*[:：]\s*(.+)$", line)
        if m:
            flush(); name, r = extract_r(m.group(1))
            cur = {"feature": feature, "name": name, "r": r, "given": [], "when": [], "then": []}
            last = cur["then"]; continue
        if cur is None: continue
        hit = None
        for kw in GIVEN:
            if line.startswith(kw): hit = ("given", line[len(kw):].strip()); break
        if not hit:
            for kw in WHEN:
                if line.startswith(kw): hit = ("when", line[len(kw):].strip()); break
        if not hit:
            for kw in THEN:
                if line.startswith(kw): hit = ("then", line[len(kw):].strip()); break
        if not hit:
            for kw in CONT:
                if line.startswith(kw):
                    if last is not None: last.append(line[len(kw):].strip())
                    hit = ("cont", None); break
        if hit and hit[0] != "cont":
            cur[hit[0]].append(hit[1]); last = cur[hit[0]]
    flush(); return cases

def title_of(md_text, fallback):
    for line in md_text.splitlines():
        m = re.match(r"^#\s+(.+)$", line.strip())
        if m: return re.split(r"[·\-—|（(]", m.group(1).strip())[0].strip()
    return fallback

# ---------- 配置表索引 + 查值(build_index / load_enums / load_keymap / lookup 见 config_index) ----------

def is_textfield(field):
    f = field.lower()
    return ("name" in f) or ("desc" in f)

def subst(text, idx, loc):
    """代入真值,返回 (新文本, [来源条目])。idx=None 则只抽引用不代值。"""
    sources = []
    def repl(m):
        ref, table, keystr, field = m.group(0), m.group(1), m.group(2).strip(), m.group(3)
        if idx is None:
            sources.append(ref); return ref
        val = lookup(idx, table, keystr, field)
        if val is None:
            sources.append(f"{ref} [查不到]"); return ref
        if val == "MULTI":
            sources.append(f"{ref} [多键/枚举·需手填]"); return ref
        cn = loc.get(val, "") if (loc and re.fullmatch(r"\d+", val) and is_textfield(field)) else ""
        shown = val + (f"（{cn}）" if cn else "")
        sources.append(f"{ref}={shown}")
        return shown
    return REF_RE.sub(repl, text), sources

# ---------- xlsx 生成 ----------
def style_header(c):
    c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER

def build(cases, title, out, idx, loc):
    join = lambda xs: "；".join(xs) if xs else "—"
    wb = Workbook()
    s = wb.active; s.title = "说明"; s.sheet_view.showGridLines = False
    s.merge_cells("A1:E1")
    s["A1"] = f"{title} · 验收测试清单(策划可读可用版)"
    s["A1"].font = Font(name=FONT, bold=True, color="FFFFFF", size=14)
    s["A1"].fill = TITLE_FILL
    s["A1"].alignment = Alignment(horizontal="center", vertical="center")
    s.row_dimensions[1].height = 30
    today = datetime.date.today().isoformat()
    note = (f"{today} 由 gherkin_to_checklist.py 从 -05验收用例.md 自动生成,共 {len(cases)} 条"
            + ("(已代入配置真值,见「来源字段」列;[需手填]/[查不到] 为护栏标记)。" if idx else "(未接配置,断言保留工程字段引用)。"))
    rows = [
        ("", ""),
        ("怎么用", "① 按「前置」把游戏调到指定状态 → ② 照「操作」点 → ③ 拿「预期结果」对实际表现核对,在「测试清单」页的 实测结果/通过 列勾选。"),
        ("真值/来源", "预期里的数字已代入配置真值;「来源字段」列标出处(`表[主键].字段`)。配置改了照来源刷新即可,逻辑不动。也可反向核对:工程/规则读表取的值与此一致否。"),
        ("关联", "每条挂工程版 R-编号,可回连 -05验收用例.md 与 -01系统规则.md。"),
        ("生成", note),
    ]
    r = 3
    for k, v in rows:
        s[f"A{r}"] = k; s[f"A{r}"].font = Font(name=FONT, bold=True, size=10, color="2F7D18")
        s.merge_cells(f"B{r}:E{r}")
        s[f"B{r}"] = v; s[f"B{r}"].font = Font(name=FONT, size=10)
        s[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="center")
        s.row_dimensions[r].height = 38; r += 1
    r += 1
    s[f"A{r}"] = "进度统计"; s[f"A{r}"].font = Font(name=FONT, bold=True, size=11, color="1D3A17"); r += 1
    last = len(cases) + 1
    stat = [("总用例数", f'=COUNTA(测试清单!A2:A{last})'),
            ("已通过",   f'=COUNTIF(测试清单!J2:J{last},"通过")'),
            ("不通过",   f'=COUNTIF(测试清单!J2:J{last},"不通过")'),
            ("待测",     "=B{0}-B{1}-B{2}"),
            ("通过率",   "=IF(B{0}=0,0,B{1}/B{0})")]
    base = r
    for i, (k, f) in enumerate(stat):
        rr = base + i
        s[f"A{rr}"] = k; s[f"A{rr}"].font = Font(name=FONT, bold=True, size=10); s[f"A{rr}"].border = BORDER
        if "{0}" in f: f = f.format(base, base+1, base+2)
        s[f"B{rr}"] = f; s[f"B{rr}"].font = Font(name=FONT, size=10); s[f"B{rr}"].border = BORDER
        s[f"B{rr}"].alignment = Alignment(horizontal="center")
    s[f"B{base+4}"].number_format = "0.0%"
    for col, w in zip("ABCDE", [12, 22, 10, 10, 30]): s.column_dimensions[col].width = w

    t = wb.create_sheet("测试清单"); t.sheet_view.showGridLines = False
    headers = ["编号", "功能", "测试点", "前置", "操作", "预期结果", "来源字段", "关联R编号", "实测结果", "通过"]
    for i, h in enumerate(headers): style_header(t.cell(row=1, column=1+i, value=h))
    t.row_dimensions[1].height = 24
    for ri, c in enumerate(cases, start=2):
        g, sg = subst(join(c["given"]), idx, loc)
        w, sw = subst(join(c["when"]),  idx, loc)
        th, st = subst(join(c["then"]), idx, loc)
        src = []
        for x in sg + sw + st:
            if x not in src: src.append(x)
        row = [f"T{ri-1:02d}", c["feature"], c["name"], g, w, th, "；".join(src) or "—", c["r"]]
        for i, v in enumerate(row):
            cc = t.cell(row=ri, column=1+i, value=v)
            cc.font = Font(name=FONT, size=10); cc.border = BORDER
            cc.alignment = Alignment(vertical="center", wrap_text=True,
                                     horizontal="center" if i in (0, 7) else "left")
        t.cell(row=ri, column=2).fill = GRP_FILL
        for col in (9, 10):
            cc = t.cell(row=ri, column=col); cc.border = BORDER
            cc.alignment = Alignment(vertical="center", horizontal="center")
    for i, w in enumerate([7, 13, 19, 24, 14, 38, 26, 15, 18, 8]):
        t.column_dimensions[chr(65+i)].width = w
    t.freeze_panes = "A2"
    t.auto_filter.ref = f"A1:J{len(cases)+1}"
    dv = DataValidation(type="list", formula1='"通过,不通过,待测"', allow_blank=True)
    t.add_data_validation(dv); dv.add(f"J2:J{len(cases)+1}")
    wb.save(out)

if __name__ == "__main__":
    args = sys.argv[1:]
    config_dir = loc_path = keymap_path = enums_path = None
    pos, i = [], 0
    while i < len(args):
        a = args[i]
        if   a == "--config" and i+1 < len(args): config_dir = args[i+1]; i += 2
        elif a == "--loc"    and i+1 < len(args): loc_path   = args[i+1]; i += 2
        elif a == "--keymap" and i+1 < len(args): keymap_path= args[i+1]; i += 2
        elif a == "--enums"  and i+1 < len(args): enums_path = args[i+1]; i += 2
        else: pos.append(a); i += 1
    if not pos:
        print("usage: python gherkin_to_checklist.py <验收用例.md> [out.xlsx] "
              "[--config <配置目录>] [--enums <枚举字典.md>] [--keymap <复合键映射.json>] [--loc <LocalizationText.xlsx>]"); sys.exit(1)
    src = pos[0]
    out = pos[1] if len(pos) > 1 else os.path.splitext(src)[0] + "-策划版.xlsx"
    with open(src, "r", encoding="utf-8") as f: md = f.read()
    cases = parse(md)
    if not cases:
        print("WARN: 没解析到场景(检查 Gherkin 『场景:』格式)"); sys.exit(2)
    idx = None
    if config_dir:
        idx = build_index(config_dir)
        if not keymap_path:                                   # 复合键映射:--keymap 优先,否则自动找配置目录里的
            auto = os.path.join(config_dir, "复合键映射.json")
            if os.path.exists(auto): keymap_path = auto
        if keymap_path:
            try: idx["__keymap__"] = load_keymap(keymap_path)
            except Exception as e: print("WARN: keymap 解析失败:", e)
        if enums_path:
            try: idx["__enums__"] = load_enums(enums_path)
            except Exception as e: print("WARN: enums 解析失败:", e)
    loc = None
    if loc_path:
        try:
            import resolve_loc
            loc = resolve_loc.build(loc_path)
        except Exception as e:
            print("WARN: --loc 解析失败,跳过文本代入:", e)
    build(cases, title_of(md, os.path.basename(src)), out, idx, loc)
    extra = ""
    if idx is not None:
        ntab = sum(1 for k in idx if not k.startswith("__"))
        flags = sum(1 for c in cases for x in
                    subst("；".join(c["given"]+c["when"]+c["then"]), idx, loc)[1] if "[" in x)
        extra = (f" | 配置表 {ntab} 张 | keymap {len(idx.get('__keymap__',{}))} 表"
                 f" | enums {len(idx.get('__enums__',{}))} 名 | 护栏标记 {flags} 处")
    print(f"saved: {out} | cases: {len(cases)}{extra}")
